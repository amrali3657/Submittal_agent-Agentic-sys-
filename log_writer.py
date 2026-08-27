"""
Opens the existing "Shop Drawing Log" workbook in-memory and updates the
"Ongoing & Submitted" sheet: appends new submittals, and updates only the
sync-owned columns (description, TUL #, Jacobs #, date submitted, status)
of rows that already exist — matched on "TUL Submittal #" (col E). Every
other column (Location, Date Received from Subs, Date Responded by Jacobs,
Date Sent to Subcontractor) is left completely untouched, since those are
filled in manually and SharePoint has no reliable source for them.
"""
import io
from datetime import datetime, date

from openpyxl import load_workbook

from config import Config


def _to_excel_date(value):
    """Graph often returns ISO datetime strings; openpyxl wants a date/datetime."""
    if not value:
        return None
    if isinstance(value, (datetime, date)):
        return value
    try:
        return datetime.fromisoformat(str(value).replace("Z", "+00:00")).date()
    except ValueError:
        return value  # leave as-is (plain string) rather than guess wrong


def update_workbook(cfg: Config, existing_bytes: bytes, submittals: list[dict]):
    if not existing_bytes:
        raise RuntimeError(
            f"No existing workbook found at {cfg.DBX_LOG_PATH} — refusing to "
            f"create one from scratch since the real log has manual columns "
            f"and a second 'Upcoming' sheet this script doesn't manage."
        )

    wb = load_workbook(io.BytesIO(existing_bytes))
    if cfg.DBX_SHEET_NAME not in wb.sheetnames:
        raise RuntimeError(
            f"Sheet '{cfg.DBX_SHEET_NAME}' not found in workbook. "
            f"Available sheets: {wb.sheetnames}"
        )
    ws = wb[cfg.DBX_SHEET_NAME]

    cols = cfg.LOG_COLUMNS
    key_col = cols["tul_submittal_no"]
    header_row = cfg.LOG_HEADER_ROW
    first_data_row = cfg.LOG_FIRST_DATA_ROW

    # Map existing TUL Submittal # -> row number
    existing_rows = {}
    row = first_data_row
    while ws[f"{key_col}{row}"].value not in (None, ""):
        val = str(ws[f"{key_col}{row}"].value).strip()
        existing_rows[val] = row
        row += 1
    next_free_row = row

    now_str = datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC")
    added, changed = [], []

    field_to_value = {
        "description": lambda s: s["description"],
        "tul_submittal_no": lambda s: s["tul_submittal_no"],
        "jacobs_submittal_no": lambda s: s["jacobs_submittal_no"],
        "date_submitted_to_jacobs": lambda s: _to_excel_date(s["start_date"]),
        "status": lambda s: s["status"],
    }

    for s in submittals:
        key = str(s["tul_submittal_no"]).strip()
        if not key:
            continue

        target_row = existing_rows.get(key)
        is_new = target_row is None
        if is_new:
            target_row = next_free_row
            existing_rows[key] = target_row
            next_free_row += 1
            added.append(key)
        else:
            current_status = ws[f'{cols["status"]}{target_row}'].value
            if str(current_status or "").strip() != str(s["status"]).strip():
                changed.append(key)

        for field in cfg.SYNCED_FIELDS:
            col_letter = cols[field]
            ws[f"{col_letter}{target_row}"] = field_to_value[field](s)

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue(), added, changed
